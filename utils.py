"""
وحدة المساعدة - توفر وظائف مفيدة مشتركة للتطبيق
"""
import io
import datetime
from flask import send_file
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def export_to_excel(data, headers, filename, sheet_name="البيانات", direction="rtl", project_summary=None):
    """
    تصدير البيانات إلى ملف إكسيل
    
    Args:
        data: البيانات المراد تصديرها (قائمة من القواميس أو قائمة من القوائم)
        headers: عناوين الأعمدة (قاموس يربط المفاتيح في البيانات بالعناوين العربية)
        filename: اسم الملف الناتج
        sheet_name: اسم ورقة العمل
        direction: اتجاه النص (rtl للعربية، ltr للإنجليزية)
        project_summary: بيانات ملخص المشروع (اختياري)
    
    Returns:
        flask.Response: استجابة فلاسك لتنزيل الملف
    """
    # إنشاء ملف إكسيل جديد
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    
    # إعداد الاتجاه المناسب للورقة
    if direction == "rtl":
        sheet.sheet_view.rightToLeft = True
    
    # إعداد نمط الخط والحدود للعناوين
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # إعداد نمط الخط للبيانات
    data_font = Font(name='Arial', size=11)
    data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # إعداد نمط خط للعناوين الفرعية
    subtitle_font = Font(name='Arial', size=12, bold=True)
    subtitle_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # إعداد نمط الحدود
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # متغير يحدد صف البداية (سيتغير إذا تم إضافة ملخص المشروع)
    start_row = 1
    
    # إضافة ملخص المشروع إذا كان موجوداً
    if project_summary:
        # عنوان التقرير
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = f"تقرير مصاريف مشروع: {project_summary.get('project_name', '')}"
        title_font = Font(name='Arial', size=14, bold=True)
        title_cell.font = title_font
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # إضافة معلومات ملخص المشروع
        summary_data = [
            ["قيمة العقد", "{:,.2f}".format(project_summary.get('contract_value', 0)), 
             "الميزانية التقديرية", "{:,.2f}".format(project_summary.get('estimated_budget', 0))],
            ["إجمالي المصاريف", "{:,.2f}".format(project_summary.get('total_expenses', 0)),
             "نسبة الإنفاق", "{:.1f}%".format(project_summary.get('expense_percentage', 0))]
        ]
        
        # إضافة الصفوف
        for i, row_data in enumerate(summary_data, 3):
            for j, value in enumerate(row_data, 1):
                cell = sheet.cell(row=i, column=j)
                cell.value = value
                cell.border = thin_border
                # تطبيق نمط خاص للعناوين
                if j % 2 == 1:  # العناوين في الأعمدة الفردية (1, 3, 5)
                    cell.font = subtitle_font
                    cell.fill = subtitle_fill
                else:  # القيم في الأعمدة الزوجية (2, 4, 6)
                    cell.font = data_font
                cell.alignment = data_alignment
        
        # إضافة ملاحظة عملة
        currency_cell = sheet.cell(row=5, column=1)
        currency_cell.value = f"جميع المبالغ بعملة: {project_summary.get('currency', 'ريال سعودي')}"
        currency_cell.font = data_font
        sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=6)
        
        # تعديل صف البداية للبيانات الرئيسية
        start_row = 7
    
    # إضافة صف العناوين للبيانات الرئيسية
    header_row = start_row
    for col_idx, (key, title) in enumerate(headers.items(), 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        cell.value = title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # إضافة البيانات
    for row_idx, row_data in enumerate(data, start_row + 1):
        for col_idx, key in enumerate(headers.keys(), 1):
            # إذا كان البيانات قائمة من القواميس
            if isinstance(row_data, dict):
                cell_value = row_data.get(key, "")
            # إذا كان البيانات قائمة من القوائم
            elif isinstance(row_data, (list, tuple)) and col_idx <= len(row_data):
                cell_value = row_data[col_idx - 1]
            else:
                cell_value = ""
            
            # معالجة خاصة للتاريخ
            if isinstance(cell_value, datetime.datetime) or isinstance(cell_value, datetime.date):
                cell_value = cell_value.strftime('%Y-%m-%d')
            
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # تعديل عرض الأعمدة ليناسب المحتوى
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        # مراجعة جميع الصفوف بما في ذلك ملخص المشروع إن وجد
        for row_idx in range(1, len(data) + start_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = max(max_length + 4, 15)  # إضافة هامش ووضع حد أدنى للعرض
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # حفظ الملف في الذاكرة
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # بناء اسم الملف كاملًا مع التاريخ
    current_date = datetime.datetime.now().strftime('%Y%m%d')
    full_filename = f"{filename}_{current_date}.xlsx"
    
    # إرسال الملف للتنزيل
    return send_file(
        output,
        as_attachment=True,
        download_name=full_filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )